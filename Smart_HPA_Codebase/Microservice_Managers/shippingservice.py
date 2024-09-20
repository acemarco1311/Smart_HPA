
#***************************************************** This file outlines the algorithm for Microservice Manager responsible for shippingservice microservice **********************************************

from Microservice_Managers.write_to_knowledge_base import write_content
import sys
import os
import fnmatch
import glob, os
import subprocess
import math
import numpy as np
import time
import statistics
from openpyxl import Workbook, load_workbook


def command_error_check(command):
    try:
        command_output = subprocess.check_output(command.split()).decode('utf-8')
        return command_output
    except:
        return None



#************************************************************************************ Monitor Component **********************************************************************************************

def Monitor ():

    microservice_name = "shippingservice"


    #                                                                   Monitoring Current Replicas and CPU Utilization per Replica

    Available_Replicas = 1
    Operational_Replicas = 0

    while Available_Replicas != Operational_Replicas and Available_Replicas is not None and Operational_Replicas is not None:               # To accomodate pod initialization time, keep trying until pod running

        Available_Replicas = "kubectl get deployment shippingservice -o=jsonpath='{.status.availableReplicas}'"
        #Available_Replicas = subprocess.check_output(Available_Replicas.split()).decode('utf-8')
        #Available_Replicas = int(Available_Replicas.strip("'"))
        Available_Replicas = command_error_check(Available_Replicas)
        if Available_Replicas is not None:
            Available_Replicas = int(Available_Replicas.strip("'"))

        Replicas_CPU_usage = 'kubectl top pods -l app=shippingservice'                                     # Monitoring CPU Usage per replica
        # Replicas_CPU_usage = subprocess.check_output(Replicas_CPU_usage.split()).decode('utf-8')
        Replicas_CPU_usage = command_error_check(Replicas_CPU_usage)
        #TODO
        if Replicas_CPU_usage is not None and len(Replicas_CPU_usage) == 0:
            Replicas_CPU_usage = None

        # error: No resources found in the namespace
        Operational_Replicas = None
        if Replicas_CPU_usage is not None:
            Operational_Replicas = len(Replicas_CPU_usage.splitlines()) - 1 # if pods running, this command reset available_replicas = operational_reps to end the while loop
        else:
            Operational_Replicas = None



        # Operational_Replicas = len(Replicas_CPU_usage.splitlines()) - 1                                         # Monitoring Operational Replicas


    current_replicas = Available_Replicas

    #                                                                   Calculating total Averaged current CPU Utilization

    cpu_add = []
    if Replicas_CPU_usage is not None:
        for line in Replicas_CPU_usage.splitlines()[1:]:
            columns = line.split()
            cpu_usage = columns[1]
            cpu_usage = cpu_usage[:-1]
            cpu_add = cpu_add + [int(cpu_usage)]
    else:
        cpu_add = None

    current_cpu = None
    if cpu_add is not None:
        current_cpu = math.ceil(statistics.mean(cpu_add))



    # cpu_add = []
    # for line in Replicas_CPU_usage.splitlines()[1:]:
    #     columns = line.split()
    #     cpu_usage = columns[1]
    #     cpu_usage = cpu_usage[:-1]
    #     cpu_add = cpu_add + [int(cpu_usage)]

    # current_cpu = math.ceil(statistics.mean(cpu_add))

    #                                                                   Monitoring Desired Replicas

    Desired_Replicas = "kubectl get deployment shippingservice -o=jsonpath='{.spec.replicas}'"
    #Desired_Replicas = subprocess.check_output(Desired_Replicas.split()).decode('utf-8')
    #Desired_Replicas = int(Desired_Replicas.strip("'"))
    Desired_Replicas = command_error_check(Desired_Replicas)
    if Desired_Replicas is not None:
        Desired_Replicas = int(Desired_Replicas.strip("'"))




    #                                                                   Monitoring CPU Request Value for a Replica Deployment

    cpu_request = "kubectl get deployment shippingservice -o=jsonpath='{.spec.template.spec.containers[0].resources.requests.cpu}'"
    #cpu_request = subprocess.check_output(cpu_request.split()).decode('utf-8')
    #cpu_request = cpu_request[:-2]
    #cpu_request = int(cpu_request.strip("'"))
    cpu_request = command_error_check(cpu_request)
    if cpu_request is not None:
        cpu_request = cpu_request[:-2]
        cpu_request = int(cpu_request.strip("'"))



    #                                                                  Setting SLA Metrics: CPU Threshold, Minimum Replicas, and Maximum Replicas

    target_cpu = 50                                        # Example Values for the Experimental Scenario 5R-50%
    max_replica = 5
    min_replica = 1

    return microservice_name, Desired_Replicas, current_replicas, current_cpu, target_cpu, cpu_request, max_replica, min_replica




#***************************************************************************** Analyze Component *****************************************************************************************


def Analyse (Desired_Replicas, current_replicas, current_cpu, target_cpu, cpu_request, min_replica):


    try:
        cpu_percentage = (current_cpu / cpu_request) * 100
        previous_desired_replicas = Desired_Replicas
        desired_replica = math.ceil(int(current_replicas) * (int(cpu_percentage) / int(target_cpu)))
        if (previous_desired_replicas != desired_replica) and (desired_replica > current_replicas) and (desired_replica >= min_replica):
            scaling_action = "scale up"
        elif (previous_desired_replicas != desired_replica) and (desired_replica < current_replicas) and (desired_replica >= min_replica):
            scaling_action = "scale down"
        else:
            scaling_action = "no scale"
        return cpu_percentage, scaling_action, desired_replica
    except:
        return None, "ERROR", None


    # cpu_percentage = (current_cpu / cpu_request) * 100                                 # Calcuating Percent CPU Utilization for the microservice deployment

    # previous_desired_replicas = Desired_Replicas                                       # To avoid making same scaling decision, when the previous scaling decision is being excecuted

    # desired_replica = math.ceil(int(current_replicas) * (int(cpu_percentage)/int(target_cpu)))                                   # Threshold-based Scaling Policy


    # # Scale up

    # if (previous_desired_replicas != desired_replica) and (desired_replica > current_replicas) and (desired_replica >= min_replica):
    #     scaling_action = "scale up"


    # # Scale down


    # elif (previous_desired_replicas != desired_replica) and (desired_replica < current_replicas) and (desired_replica >= min_replica):
    #     scaling_action = "scale down"


    # #  No Scaling

    # else:
    #     scaling_action = "no scale"



    # return cpu_percentage, scaling_action, desired_replica





#********************************************************************** Microservice_Main_Function ********************************************************************************


def shippingservice(Test_Time):

    microservice_name, P_Desired_Replicas, current_replicas, current_cpu, target_cpu, cpu_request, max_replica, min_replica = Monitor ()
    cpu_percentage, scaling_action, desired_replica = Analyse (P_Desired_Replicas, current_replicas, current_cpu, target_cpu, cpu_request, min_replica)

    shippingservice_data = [microservice_name, scaling_action, desired_replica, current_replicas, cpu_request, max_replica]


    #********************************************************************** Storing Data for each iteration in the Knowledge Base ********************************************************************************

    #workbook = load_workbook('./Knowledge_Base/shippingservice.xlsx')
    #sheet = workbook.active

    #sheet.cell(row=1, column=1, value="Test Time (sec)")
    #sheet.cell(row=1, column=2, value="CPU Usage Percentage")
    #sheet.cell(row=1, column=3, value="Current Replicas")
    #sheet.cell(row=1, column=4, value="Desired Replicas")
    #sheet.cell(row=1, column=5, value="Max. Replicas")
    #sheet.cell(row=1, column=6, value="Scaling Action")



    #Test_Time_row_length = len(sheet['A']) + 1
    #CPU_usage_row_length = len(sheet['B']) + 1
    #current_replica_row_length = len(sheet['C']) + 1
    #desired_replica_row_length = len(sheet['D']) + 1



    #sheet[f'A{Test_Time_row_length}'] = Test_Time
    #sheet[f'B{CPU_usage_row_length}'] = cpu_percentage
    #sheet[f'C{current_replica_row_length}'] = current_replicas
    #sheet[f'D{desired_replica_row_length}'] = desired_replica


    #workbook.save('./Knowledge_Base/shippingservice.xlsx')

    write_content('./Knowledge_Base/shippingservice.txt', Test_Time, cpu_percentage, current_replicas, desired_replica, max_replica, scaling_action)



    return shippingservice_data


